import openpyxl

wb = openpyxl.load_workbook('(Apr to Oct) 202526Sales.XLSX', read_only=True, data_only=True)
print("Sheets:", wb.sheetnames)
ws = wb.active

# Print first 10 rows to understand structure
print("\nFirst 10 rows:")
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
    print(f"Row {i+1}: {[str(v or '')[:30] for v in row[:15]]}")

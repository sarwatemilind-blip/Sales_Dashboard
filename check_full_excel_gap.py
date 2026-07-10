import openpyxl
import csv
from collections import defaultdict
from datetime import datetime

print("Reading Excel file...")
wb = openpyxl.load_workbook('(Apr to Oct) 202526Sales.XLSX', read_only=True, data_only=True)
ws = wb.active

# Row 3 = headers, data from row 4
headers = None
for i, row in enumerate(ws.iter_rows(min_row=1, max_row=3, values_only=True)):
    if i == 2:  # row 3 (0-indexed = 2)
        headers = [str(v or '').strip() for v in row]

print("Headers:", headers[:20])

# Load stockist_mapping codes
mapping_codes = set()
with open('stockist_mapping.csv', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        mapping_codes.add(row['stockist_code'].strip())

month_labels = {1:'Apr', 2:'May', 3:'Jun', 4:'Jul', 5:'Aug', 6:'Sep',
                7:'Oct', 8:'Nov', 9:'Dec', 10:'Jan', 11:'Feb', 12:'Mar'}

def get_fiscal_month(date_obj):
    m = date_obj.month
    return m - 3 if m >= 4 else m + 9

def get_fiscal_year(date_obj):
    return date_obj.year if date_obj.month >= 4 else date_obj.year - 1

total_by_month = defaultdict(float)
missing_by_month = defaultdict(lambda: defaultdict(float))
missing_info = {}
row_count = 0
error_rows = 0

for row in ws.iter_rows(min_row=4, values_only=True):
    vals = list(row)
    if not any(v for v in vals[:5]):
        continue
    d = dict(zip(headers, [str(v or '').strip() for v in vals]))

    # Get bill date to determine fiscal month/year
    bill_date_raw = vals[5] if len(vals) > 5 else None  # Bill Date is col 6 (index 5)
    try:
        if isinstance(bill_date_raw, datetime):
            bd = bill_date_raw
        elif bill_date_raw:
            bd = datetime.strptime(str(bill_date_raw)[:10], '%Y-%m-%d')
        else:
            error_rows += 1
            continue
        fy = get_fiscal_year(bd)
        fm = get_fiscal_month(bd)
    except:
        error_rows += 1
        continue

    # Get amount - look for Net Value or similar
    amt_raw = None
    for col_name in ['Net Value', 'Net Amt', 'Amount', 'Value', 'Net']:
        for k, v in d.items():
            if col_name.lower() in k.lower() and v:
                try:
                    amt_raw = float(v)
                    break
                except:
                    pass
        if amt_raw is not None:
            break

    if amt_raw is None:
        # Try last numeric columns
        for v in reversed(vals):
            if v and isinstance(v, (int, float)):
                amt_raw = float(v)
                break

    if amt_raw is None:
        error_rows += 1
        continue

    sc = d.get('Stockist Code', '').strip()
    sname = d.get('Stockist Name', '').strip()
    hq_code = d.get('HQ Code', '').strip()
    hq_name = d.get('HQ', '').strip()
    dist = d.get('Distributor Code', '').strip()

    key = (fy, fm)
    total_by_month[key] += amt_raw
    row_count += 1

    if sc and sc not in mapping_codes:
        missing_by_month[key][sc] += amt_raw
        if sc not in missing_info:
            missing_info[sc] = {'name': sname, 'hq_code': hq_code, 'hq_name': hq_name, 'dist': dist}

print(f"Processed: {row_count} rows, {error_rows} skipped")
print(f"\n{'Month':<12} {'Total (L)':>12} {'Gap (L)':>10} {'% Gap':>7}")
print("-"*47)
grand_total = grand_gap = 0
for (yr, mo) in sorted(total_by_month.keys()):
    t = total_by_month[(yr,mo)]
    g = sum(missing_by_month[(yr,mo)].values())
    label = f"FY{yr} {month_labels.get(mo,mo)}"
    print(f"{label:<12} {t/100000:>12.2f} {g/100000:>10.2f} {(g/t*100 if t else 0):>6.1f}%")
    grand_total += t
    grand_gap += g

print("-"*47)
print(f"{'TOTAL':<12} {grand_total/100000:>12.2f} {grand_gap/100000:>10.2f} {(grand_gap/grand_total*100 if grand_total else 0):>6.1f}%")

if missing_info:
    all_totals = defaultdict(float)
    for mo_data in missing_by_month.values():
        for sc, amt in mo_data.items():
            all_totals[sc] += amt

    print(f"\nMissing stockists ({len(missing_info)}):")
    print(f"{'Code':<8} {'Name':<40} {'HQ':<8} {'Total (L)':>10}")
    print("-"*70)
    for sc, tot in sorted(all_totals.items(), key=lambda x: -x[1]):
        info = missing_info.get(sc, {})
        print(f"{sc:<8} {info.get('name','')[:39]:<40} {info.get('hq_code',''):<8} {tot/100000:>10.2f}")
else:
    print("\nNo missing stockists!")

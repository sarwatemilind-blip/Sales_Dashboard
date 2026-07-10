import openpyxl

wb = openpyxl.load_workbook('CFA-Stockist-HQ_Mapping_Master_Updated.xlsx', read_only=True, data_only=True)
ws = wb.active

headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
print("All headers:", headers)
print()

# HQ codes we need to map: S0154, S0270, S0142, S0359, S0180, S0227, S0073
target_hqs = {'S0154','S0270','S0142','S0359','S0180','S0227','S0073'}

# Find all stockists in these HQs to guess BE assignments
from collections import defaultdict
hq_bes = defaultdict(set)

for row in ws.iter_rows(min_row=2, values_only=True):
    vals = [str(v or '').strip() for v in row]
    d = dict(zip(headers, vals))
    hq = d.get('HQ Code','').strip()
    if hq in target_hqs:
        be1 = d.get('BE Id - 1','').strip()
        be2 = d.get('BE Id - 2','').strip()
        be3 = d.get('BE Id - 3','').strip()
        asm = d.get('ASM EMP id','').strip()
        rsm = d.get('RSM Code','').strip()
        # Get all remaining cols for ZM, VP
        all_ids = [be1, be2, be3, asm, rsm]
        extra_vals = list(d.values())[len(headers)-5:] if len(d) > len(headers)-5 else []
        
        row_str = f"BE1={be1} BE2={be2} BE3={be3} ASM={asm} RSM={rsm}"
        hq_bes[hq].add(row_str)

for hq in sorted(target_hqs):
    print(f"\nHQ {hq} - Existing BE assignments:")
    for r in list(hq_bes[hq])[:5]:
        print(f"  {r}")

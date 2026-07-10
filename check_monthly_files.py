import openpyxl
import csv
from collections import defaultdict
from datetime import datetime

files = [
    'Mar26 Sales.XLSX',
    'Feb26 Sales.XLSX', 
    'Jan26 Sales.XLSX',
    'Nov25 Sales.XLSX',
    'Dec25 Sales.XLSX',
]

# Load stockist_mapping codes
mapping_codes = set()
with open('stockist_mapping.csv', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        mapping_codes.add(row['stockist_code'].strip())

month_labels = {1:'Apr', 2:'May', 3:'Jun', 4:'Jul', 5:'Aug', 6:'Sep',
                7:'Oct', 8:'Nov', 9:'Dec', 10:'Jan', 11:'Feb', 12:'Mar'}

def get_fiscal_month(date_obj):
    m = date_obj.month
    return m - 3 if m >= 4 else m + 9

def get_fiscal_year(date_obj):
    return date_obj.year if date_obj.month >= 4 else date_obj.year - 1

total_by_month = defaultdict(float)
missing_by_month = defaultdict(lambda: defaultdict(float))
missing_info = {}

for fname in files:
    print(f"\nReading {fname}...")
    try:
        wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
    except Exception as e:
        print(f"  Error: {e}")
        continue
    ws = wb.active

    # Find header row
    headers = None
    header_row = 0
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True)):
        row_strs = [str(v or '').strip() for v in row]
        if any('stockist' in v.lower() or 'bill date' in v.lower() or 'distributor code' in v.lower() for v in row_strs):
            headers = row_strs
            header_row = i + 1
            break

    if not headers:
        print(f"  Could not find header row!")
        continue
    
    print(f"  Header at row {header_row}: {headers[:10]}")

    row_count = 0
    for row in ws.iter_rows(min_row=header_row+1, values_only=True):
        vals = list(row)
        if not any(v for v in vals[:5]):
            continue
        d = dict(zip(headers, [str(v or '').strip() for v in vals]))

        # Bill Date
        bill_date_raw = vals[header_row-1] if False else None
        for idx, h in enumerate(headers):
            if 'bill date' in h.lower() or h.lower() == 'date':
                bill_date_raw = vals[idx]
                break

        try:
            if isinstance(bill_date_raw, datetime):
                bd = bill_date_raw
            elif bill_date_raw:
                bd = datetime.strptime(str(bill_date_raw)[:10], '%Y-%m-%d')
            else:
                continue
            fy = get_fiscal_year(bd)
            fm = get_fiscal_month(bd)
        except:
            continue

        # Amount
        amt_raw = None
        for idx, h in enumerate(headers):
            if any(kw in h.lower() for kw in ['net value', 'net amt', 'amount', 'value', 'net']):
                try:
                    v = vals[idx]
                    if v and isinstance(v, (int, float)):
                        amt_raw = float(v)
                        break
                    elif v:
                        amt_raw = float(str(v).replace(',',''))
                        break
                except:
                    pass

        if amt_raw is None:
            for v in reversed(vals):
                if v and isinstance(v, (int, float)) and float(v) > 0:
                    amt_raw = float(v)
                    break

        if amt_raw is None:
            continue

        sc = d.get('Stockist Code', '').strip()
        key = (fy, fm)
        total_by_month[key] += amt_raw
        row_count += 1

        if sc and sc not in mapping_codes:
            missing_by_month[key][sc] += amt_raw
            if sc not in missing_info:
                missing_info[sc] = {
                    'name': d.get('Stockist Name', ''),
                    'hq_code': d.get('HQ Code', ''),
                    'hq_name': d.get('HQ', ''),
                    'dist': d.get('Distributor Code', ''),
                }
    print(f"  Processed {row_count} rows")

print(f"\n{'Month':<12} {'Total (L)':>12} {'Gap (L)':>10} {'% Gap':>7}")
print("-"*47)
grand_total = grand_gap = 0
for (yr, mo) in sorted(total_by_month.keys()):
    t = total_by_month[(yr,mo)]
    g = sum(missing_by_month[(yr,mo)].values())
    label = f"FY{yr} {month_labels.get(mo,mo)}"
    print(f"{label:<12} {t/100000:>12.2f} {g/100000:>10.2f} {(g/t*100 if t else 0):>6.1f}%")
    grand_total += t
    grand_gap += g
print("-"*47)
print(f"{'TOTAL':<12} {grand_total/100000:>12.2f} {grand_gap/100000:>10.2f} {(grand_gap/grand_total*100 if grand_total else 0):>6.1f}%")

if missing_info:
    all_totals = defaultdict(float)
    for mo_data in missing_by_month.values():
        for sc, amt in mo_data.items():
            all_totals[sc] += amt
    print(f"\nMissing stockists ({len(missing_info)}):")
    print(f"{'Code':<8} {'Name':<40} {'HQ':<8} {'Total (L)':>10}")
    print("-"*70)
    for sc, tot in sorted(all_totals.items(), key=lambda x: -x[1]):
        info = missing_info.get(sc, {})
        print(f"{sc:<8} {info.get('name','')[:39]:<40} {info.get('hq_code',''):<8} {tot/100000:>10.2f}")
else:
    print("\nNo missing stockists!")

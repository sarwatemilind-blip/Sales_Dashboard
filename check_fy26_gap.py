import openpyxl
import csv
from collections import defaultdict
from datetime import datetime

files = [
    'APR26 SALE.XLSX',
    'May26 Sales Interact.XLSX',
    'June 2026 Sales.XLSX',
    'Jul Sale_07_07_2026.XLSX',
]

# Load stockist_mapping codes
mapping_codes = set()
with open('stockist_mapping.csv', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        mapping_codes.add(row['stockist_code'].strip())

month_labels = {1:'Apr-26', 2:'May-26', 3:'Jun-26', 4:'Jul-26', 5:'Aug-26', 6:'Sep-26',
                7:'Oct-26', 8:'Nov-26', 9:'Dec-26', 10:'Jan-27', 11:'Feb-27', 12:'Mar-27'}

def get_fiscal_month(date_obj):
    m = date_obj.month
    return m - 3 if m >= 4 else m + 9

def get_fiscal_year(date_obj):
    return date_obj.year if date_obj.month >= 4 else date_obj.year - 1

total_by_month = defaultdict(float)
missing_by_month = defaultdict(lambda: defaultdict(float))
missing_info = {}

for fname in files:
    print(f"\nReading {fname}...")
    try:
        wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
    except Exception as e:
        print(f"  Error: {e}")
        continue
    ws = wb.active

    # Find header row (look for Stockist Code or Bill Date)
    headers = None
    header_row_num = 0
    all_rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
    for i, row in enumerate(all_rows):
        row_strs = [str(v or '').strip().lower() for v in row]
        if any('stockist' in v or 'bill date' in v or 'billno' in v for v in row_strs):
            headers = [str(v or '').strip() for v in row]
            header_row_num = i + 1
            break

    if not headers:
        print(f"  ERROR: Could not find header row")
        # Print first 5 rows for debug
        for i, row in enumerate(all_rows[:5]):
            print(f"  Row {i+1}: {[str(v or '')[:20] for v in row[:10]]}")
        continue

    print(f"  Headers (row {header_row_num}): {headers[:12]}")

    # Find key column indices
    def find_col(keywords):
        for idx, h in enumerate(headers):
            hl = h.lower()
            if any(kw in hl for kw in keywords):
                return idx
        return None

    date_col = find_col(['bill date', 'date'])
    sc_col = find_col(['stockist code', 'stockist_code'])
    sname_col = find_col(['stockist name', 'stockist_name'])
    hq_col = find_col(['hq code', 'hq_code'])
    hqname_col = find_col(['hq name', "'hq'"])
    dist_col = find_col(['distributor code', 'distributor_code'])

    # For amount - find rightmost numeric-looking column
    amt_col = find_col(['net value', 'net amt', 'net_value', 'amount', 'value'])

    print(f"  date_col={date_col}, sc_col={sc_col}, amt_col={amt_col}")

    row_count = 0
    for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
        vals = list(row)
        if not any(v for v in vals[:5]):
            continue

        # Bill date
        bill_date_raw = vals[date_col] if date_col is not None and date_col < len(vals) else None
        try:
            if isinstance(bill_date_raw, datetime):
                bd = bill_date_raw
            elif bill_date_raw:
                bd = datetime.strptime(str(bill_date_raw)[:10], '%Y-%m-%d')
            else:
                continue
            fy = get_fiscal_year(bd)
            fm = get_fiscal_month(bd)
        except:
            continue

        # Amount
        amt_raw = None
        if amt_col is not None and amt_col < len(vals):
            try:
                v = vals[amt_col]
                if v is not None:
                    amt_raw = float(str(v).replace(',', ''))
            except:
                pass
        if amt_raw is None:
            # Try last numeric value in row
            for v in reversed(vals):
                if isinstance(v, (int, float)) and v != 0:
                    amt_raw = float(v)
                    break
        if amt_raw is None:
            continue

        sc = str(vals[sc_col] or '').strip() if sc_col is not None and sc_col < len(vals) else ''
        sname = str(vals[sname_col] or '').strip() if sname_col is not None else ''
        hq_code = str(vals[hq_col] or '').strip() if hq_col is not None else ''
        hq_name = str(vals[hqname_col] or '').strip() if hqname_col is not None else ''
        dist = str(vals[dist_col] or '').strip() if dist_col is not None else ''

        key = (fy, fm)
        total_by_month[key] += amt_raw
        row_count += 1

        if sc and sc not in mapping_codes:
            missing_by_month[key][sc] += amt_raw
            if sc not in missing_info:
                missing_info[sc] = {'name': sname, 'hq_code': hq_code, 'hq_name': hq_name, 'dist': dist}

    print(f"  Processed {row_count} rows")

print(f"\n\n{'Month':<12} {'Total (L)':>12} {'Gap (L)':>10} {'% Gap':>7}")
print("-"*47)
grand_total = grand_gap = 0
for (yr, mo) in sorted(total_by_month.keys()):
    t = total_by_month[(yr,mo)]
    g = sum(missing_by_month[(yr,mo)].values())
    label = month_labels.get(mo, f"M{mo}")
    print(f"{label:<12} {t/100000:>12.2f} {g/100000:>10.2f} {(g/t*100 if t else 0):>6.1f}%")
    grand_total += t
    grand_gap += g
print("-"*47)
print(f"{'TOTAL':<12} {grand_total/100000:>12.2f} {grand_gap/100000:>10.2f} {(grand_gap/grand_total*100 if grand_total else 0):>6.1f}%")

if missing_info:
    all_totals = defaultdict(float)
    for mo_data in missing_by_month.values():
        for sc, amt in mo_data.items():
            all_totals[sc] += amt

    print(f"\nMissing stockists ({len(missing_info)}) - need to add to mapping:")
    print(f"{'Code':<8} {'Name':<42} {'HQ':<8} {'Dist':<6} {'Total (L)':>10}")
    print("-"*78)
    for sc, tot in sorted(all_totals.items(), key=lambda x: -x[1]):
        info = missing_info.get(sc, {})
        print(f"{sc:<8} {info.get('name','')[:41]:<42} {info.get('hq_code',''):<8} {info.get('dist',''):<6} {tot/100000:>10.2f}")

    # Also show per-month breakdown for each missing stockist
    print(f"\nPer-month breakdown of missing stockists:")
    for sc in sorted(all_totals.keys(), key=lambda x: -all_totals[x]):
        info = missing_info.get(sc, {})
        monthly = []
        for (yr, mo) in sorted(missing_by_month.keys()):
            if missing_by_month[(yr,mo)].get(sc, 0) > 0:
                monthly.append(f"{month_labels.get(mo,'M'+str(mo))}:{missing_by_month[(yr,mo)][sc]/100000:.2f}L")
        print(f"  {sc} ({info.get('name','')[:30]}): {', '.join(monthly)}")
else:
    print("\nNo missing stockists - mapping is complete!")

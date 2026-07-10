import csv
import openpyxl
from collections import defaultdict
from datetime import datetime

# Load stockist mapping - find ones with NO BEs
no_be_stockists = set()
with open('stockist_mapping.csv', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        be1 = (row.get('be_emp_id_1') or '').strip()
        be2 = (row.get('be_emp_id_2') or '').strip()
        be3 = (row.get('be_emp_id_3') or '').strip()
        if not be1 and not be2 and not be3:
            no_be_stockists.add(row['stockist_code'].strip())

print(f"Stockists with NO BE assigned: {len(no_be_stockists)}")
for s in sorted(no_be_stockists):
    print(f"  {s}")

# Now check sales for these stockists across all 4 months
def get_fiscal_month(d): return d.month - 3 if d.month >= 4 else d.month + 9
def get_fiscal_year(d): return d.year if d.month >= 4 else d.year - 1

month_labels = {1:'Apr-26', 2:'May-26', 3:'Jun-26', 4:'Jul-26'}
files = [
    ('APR26 SALE.XLSX', 3),
    ('May26 Sales Interact.XLSX', 4),
    ('June 2026 Sales.XLSX', 3),
    ('Jul Sale_07_07_2026.XLSX', 3),
]

no_be_sales = defaultdict(lambda: defaultdict(float))  # month -> stockist -> amt

for fname, hrow in files:
    wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c or '').strip() for c in list(ws.iter_rows(min_row=hrow, max_row=hrow, values_only=True))[0]]
    date_col = next((i for i,h in enumerate(headers) if 'bill date' in h.lower()), None)
    sc_col = next((i for i,h in enumerate(headers) if 'stockist code' in h.lower()), None)
    amt_col = 49
    for row in ws.iter_rows(min_row=hrow+1, values_only=True):
        vals = list(row)
        if not any(vals[:5]): continue
        try:
            sc = str(vals[sc_col] or '').strip()
            if sc not in no_be_stockists: continue
            bd = vals[date_col]
            if not isinstance(bd, datetime): bd = datetime.strptime(str(bd)[:10], '%Y-%m-%d')
            fm = get_fiscal_month(bd)
            amt = float(vals[amt_col] or 0)
            no_be_sales[(get_fiscal_year(bd), fm)][sc] += amt
        except: pass
    wb.close()

print(f"\nSales attributed to no-BE stockists (INVISIBLE in dashboard):")
grand = 0
for (yr, mo) in sorted(no_be_sales.keys()):
    month_total = sum(no_be_sales[(yr,mo)].values())
    grand += month_total
    label = month_labels.get(mo, mo)
    print(f"\n  {label}: {month_total/100000:.2f} L")
    for sc, amt in sorted(no_be_sales[(yr,mo)].items(), key=lambda x: -x[1]):
        print(f"    {sc}: {amt/100000:.4f} L")

print(f"\nTotal invisible sales: {grand/100000:.2f} L")

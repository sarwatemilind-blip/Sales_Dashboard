import csv
import openpyxl
from collections import defaultdict
from datetime import datetime

# Load stockist mapping - find ones with NO BEs
no_be_map = {}
with open('stockist_mapping.csv', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        be1 = (row.get('be_emp_id_1') or '').strip()
        be2 = (row.get('be_emp_id_2') or '').strip()
        be3 = (row.get('be_emp_id_3') or '').strip()
        if not be1 and not be2 and not be3:
            no_be_map[row['stockist_code'].strip()] = dict(row)

def get_fiscal_month(d): return d.month - 3 if d.month >= 4 else d.month + 9
def get_fiscal_year(d): return d.year if d.month >= 4 else d.year - 1

files = [
    ('APR26 SALE.XLSX', 3),
    ('May26 Sales Interact.XLSX', 4),
    ('June 2026 Sales.XLSX', 3),
    ('Jul Sale_07_07_2026.XLSX', 3),
]

# Find which no-BE stockists actually have POSITIVE sales
sc_totals = defaultdict(float)
for fname, hrow in files:
    wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c or '').strip() for c in list(ws.iter_rows(min_row=hrow, max_row=hrow, values_only=True))[0]]
    date_col = next((i for i,h in enumerate(headers) if 'bill date' in h.lower()), None)
    sc_col = next((i for i,h in enumerate(headers) if 'stockist code' in h.lower()), None)
    sname_col = next((i for i,h in enumerate(headers) if 'stockist name' in h.lower()), None)
    for row in ws.iter_rows(min_row=hrow+1, values_only=True):
        vals = list(row)
        if not any(vals[:5]): continue
        try:
            sc = str(vals[sc_col] or '').strip()
            if sc not in no_be_map: continue
            amt = float(vals[49] or 0)
            sc_totals[sc] += amt
        except: pass
    wb.close()

# Only stockists with positive net sales
active_no_be = {sc: amt for sc, amt in sc_totals.items() if amt > 0}
print(f"No-BE stockists with positive sales: {len(active_no_be)}")
print(f"{'Code':<8} {'Name':<40} {'HQ':<8} {'Sales (L)':>10}")
print("-"*70)
for sc, amt in sorted(active_no_be.items(), key=lambda x: -x[1]):
    info = no_be_map.get(sc, {})
    print(f"{sc:<8} {info.get('stockist_name','')[:39]:<40} {info.get('hq_code',''):<8} {amt/100000:>10.4f}")

# Now look up these in master Excel
target_codes = set(active_no_be.keys())
target_hqs = {no_be_map[sc].get('hq_code','') for sc in target_codes}

print(f"\nLooking up in master Excel...")
wb = openpyxl.load_workbook('CFA-Stockist-HQ_Mapping_Master_Updated.xlsx', read_only=True, data_only=True)
ws = wb.active
headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

found_in_master = {}
hq_context = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    vals = [str(v or '').strip() for v in row]
    d = dict(zip(headers, vals))
    sc = d.get('Stockist Code','')
    hq = d.get('HQ Code','')
    if sc in target_codes:
        found_in_master[sc] = d
    if hq in target_hqs and hq not in hq_context:
        hq_context[hq] = d

print(f"\nFound {len(found_in_master)} in master Excel:")
for sc, d in sorted(found_in_master.items()):
    print(f"  {sc} | BE1:{d.get('BE Id - 1')} BE2:{d.get('BE Id - 2')} ASM:{d.get('ASM EMP id')} RSM:{d.get('RSM Code')} ZM:{d.get('ZM Code')} VP:{d.get('VP Code')}")

not_found = target_codes - set(found_in_master.keys())
print(f"\nNOT in master ({len(not_found)}) - using HQ context:")
for sc in sorted(not_found, key=lambda x: -active_no_be[x]):
    info = no_be_map.get(sc, {})
    hq = info.get('hq_code','')
    ctx = hq_context.get(hq, {})
    print(f"  {sc} ({info.get('stockist_name','')[:30]}) HQ:{hq}")
    if ctx:
        print(f"    -> BE1:{ctx.get('BE Id - 1')} BE2:{ctx.get('BE Id - 2')} ASM:{ctx.get('ASM EMP id')} RSM:{ctx.get('RSM Code')}")
    else:
        print(f"    -> HQ not in master!")

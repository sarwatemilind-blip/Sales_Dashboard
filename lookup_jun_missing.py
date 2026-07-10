import openpyxl
import csv
from collections import defaultdict

# Missing stockists info from sales
missing = {
    'C1124': {'name': 'MAA DRUGS DISTRIBUTOR',    'hq_code': 'S0188', 'dist': '021'},
    'C1125': {'name': 'JYOTI MEDICAL AGENCIES',   'hq_code': 'S0372', 'dist': '016'},
    'C0017': {'name': 'SRI MAHAVEER DRUGS',        'hq_code': 'S0148', 'dist': '014'},
    'C1127': {'name': 'NEW MEHTA BROTHERS',        'hq_code': 'S0350', 'dist': '011'},
    'C1126': {'name': 'VIKKY PHARMA',              'hq_code': 'S0037', 'dist': '011'},
}
target_hqs = {v['hq_code'] for v in missing.values()}

# Read master Excel for HQ context
wb = openpyxl.load_workbook('CFA-Stockist-HQ_Mapping_Master_Updated.xlsx', read_only=True, data_only=True)
ws = wb.active
headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

hq_context = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    vals = [str(v or '').strip() for v in row]
    d = dict(zip(headers, vals))
    hq = d.get('HQ Code','').strip()
    if hq in target_hqs and hq not in hq_context:
        hq_context[hq] = d

# Also check if stockist codes exist in master
master_stockists = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    vals = [str(v or '').strip() for v in row]
    d = dict(zip(headers, vals))
    sc = d.get('Stockist Code','').strip()
    if sc in missing:
        master_stockists[sc] = d

print("=== Stockists found in master Excel ===")
for sc, d in master_stockists.items():
    print(f"\n{sc} | {d.get('Stockist Name')}")
    print(f"  BE1:{d.get('BE Id - 1')} BE2:{d.get('BE Id - 2')} BE3:{d.get('BE Id - 3')}")
    print(f"  ASM:{d.get('ASM EMP id')} RSM:{d.get('RSM Code')} ZM:{d.get('ZM Code')} VP:{d.get('VP Code')}")
    print(f"  Area:{d.get('Area Name')} Region:{d.get('Region Name')} Zone:{d.get('Zone Name')}")

not_in_master = set(missing.keys()) - set(master_stockists.keys())
print(f"\n=== NOT in master Excel: {not_in_master} ===")
print("(Will use HQ context to infer BE assignments)\n")

for sc in not_in_master:
    info = missing[sc]
    hq = info['hq_code']
    ctx = hq_context.get(hq, {})
    print(f"\n{sc} | {info['name']} | HQ:{hq}")
    if ctx:
        print(f"  Dist: {ctx.get('Distributor Code')} | {ctx.get('Distributorname')} | {ctx.get('Distributorcity')}")
        print(f"  BE1:{ctx.get('BE Id - 1')} BE2:{ctx.get('BE Id - 2')} BE3:{ctx.get('BE Id - 3')}")
        print(f"  ASM:{ctx.get('ASM EMP id')} RSM:{ctx.get('RSM Code')} ZM:{ctx.get('ZM Code')} VP:{ctx.get('VP Code')}")
        print(f"  Area:{ctx.get('Area Name')} Region:{ctx.get('Region Name')} Zone:{ctx.get('Zone Name')}")
    else:
        print(f"  HQ {hq} NOT FOUND in master Excel!")

import openpyxl
import csv

wb = openpyxl.load_workbook('CFA-Stockist-HQ_Mapping_Master_Updated.xlsx', read_only=True, data_only=True)
ws = wb.active
headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

# HQ codes for missing stockists
target_hqs = {'S0154','S0270','S0142','S0359','S0180','S0227','S0073'}
hq_context = {}  # hq_code -> sample row with full details

for row in ws.iter_rows(min_row=2, values_only=True):
    vals = [str(v or '').strip() for v in row]
    d = dict(zip(headers, vals))
    hq = d.get('HQ Code','').strip()
    if hq in target_hqs and hq not in hq_context:
        hq_context[hq] = d

# Print context for each HQ
for hq, d in sorted(hq_context.items()):
    print(f"\n=== HQ {hq} ({d.get('HQ','')}) ===")
    print(f"  Dist: {d.get('Distributor Code')} | {d.get('Distributorname')} | {d.get('Distributorcity')}")
    print(f"  Area: {d.get('Area Name')} | Region: {d.get('Region Name')} | Zone: {d.get('Zone Name')}")
    print(f"  BE1: {d.get('BE Id - 1')} | BE2: {d.get('BE Id - 2')} | BE3: {d.get('BE Id - 3')}")
    print(f"  ASM: {d.get('ASM EMP id')} | RSM: {d.get('RSM Code')} | ZM: {d.get('ZM Code')} | VP: {d.get('VP Code')}")

# Also load sales data for the missing stockists
missing_codes = {'C1120','C1122','C1123','C1121','C0426','C1080','C1097'}
seen = {}
with open('sales.csv', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        sc = row.get('stockist_code','').strip()
        if sc in missing_codes and sc not in seen:
            seen[sc] = {
                'stockist_code': sc,
                'stockist_name': row.get('stockist_name','').strip(),
                'distributor_code': row.get('distributor_code','').strip(),
                'hq_code': row.get('hq_code','').strip(),
                'hq_name': row.get('hq_name','').strip(),
            }

print("\n\n=== Missing stockists from sales.csv ===")
for sc, d in sorted(seen.items()):
    hq = d['hq_code']
    ctx = hq_context.get(hq, {})
    print(f"\n{sc} | {d['stockist_name']} | HQ: {hq} ({d['hq_name']}) | Dist: {d['distributor_code']}")
    if ctx:
        print(f"  -> Area: {ctx.get('Area Name')} | Region: {ctx.get('Region Name')} | Zone: {ctx.get('Zone Name')}")
        print(f"  -> BE1: {ctx.get('BE Id - 1')} | BE2: {ctx.get('BE Id - 2')} | ASM: {ctx.get('ASM EMP id')} | RSM: {ctx.get('RSM Code')} | ZM: {ctx.get('ZM Code')} | VP: {ctx.get('VP Code')}")
    else:
        print(f"  -> HQ NOT in master Excel!")

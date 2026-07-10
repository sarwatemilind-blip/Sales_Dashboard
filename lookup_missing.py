import openpyxl

wb = openpyxl.load_workbook('CFA-Stockist-HQ_Mapping_Master_Updated.xlsx', read_only=True, data_only=True)
ws = wb.active

headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]

targets = {'C1120','C1122','C1123','C1121','C0426','C1080','C1097'}
found = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    vals = [str(v or '').strip() for v in row]
    d = dict(zip(headers, vals))
    sc = d.get('Stockist Code','').strip()
    if sc in targets:
        found[sc] = d
        print(f"Code: {sc}")
        print(f"  Name: {d.get('Stockist Name')}")
        print(f"  Dist: {d.get('Distributor Code')} | HQ Code: {d.get('HQ Code')} | HQ: {d.get('HQ')}")
        print(f"  BE1: {d.get('BE Id - 1')} | BE2: {d.get('BE Id - 2')} | BE3: {d.get('BE Id - 3')}")
        print(f"  ASM: {d.get('ASM EMP id')} | RSM: {d.get('RSM Code')}")
        # Print all remaining columns
        for k, v in d.items():
            if k not in ['Distributor Code','Distributorname','Distributorcity','Stockist Code','Stockist Name','Gst No','Place Of Supply','HQ Code','HQ','BE Id - 1','BE Id - 2','BE Id - 3','Name BE - 1','Name BE - 2','Name BE - 3','ASM EMP id','ASM Name','ASM HQ Code','ASM HQ','RSM Code']:
                if v and v != '':
                    print(f"  {k}: {v}")
        print()

not_found = targets - set(found.keys())
if not_found:
    print(f"NOT found in Excel: {not_found}")
